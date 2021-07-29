from openpyxl import Workbook, load_workbook
import difflib
from openpyxl.styles import PatternFill

def string_similar(s1, s2):
    '''This function takes two strings, s1 and s2 and calculates the percentage of similarity of the two strings,
    a number between 0 and 1 is the return value'''
    return difflib.SequenceMatcher(None, s1, s2).quick_ratio()

# price list is imported as excel workbook; file pathway must be specified
workbook = load_workbook('/Users/lucyliu/Downloads/FridayGrocer - Master.xlsx')
worksheet = workbook.active

# a new excel file created to compare data
new_workbook = Workbook()
worksheet_2 = new_workbook.active
worksheet_2.title = 'Data'

# column titles in a worksheet in the new excel file generated
worksheet_2.append(['Similarity Value', 'Woolworths items', 'Coles items', 'Woolworths Price', 'Coles Price', 'Price Similarity'])

# rows used defined as all rows with data except header
rows = worksheet.iter_rows(min_row=2)

# take each row in rows, and clean data for comparison
for row in rows:
    if row[4].value != None and row[7].value != None:
        row[4].value.strip()
        row[7].value.strip()
    if row[5].value != None and row[8].value != None and row[5].value != '' and row[8].value != '':
        str(row[5].value).strip('$')
        str(row[8].value).strip('$')

    # check if a valid comparison can be made(both coles/woolworths hold the item)
    if row[4].value != None and row[7].value != None and row[5].value != None and row[8].value != None and row[
        5].value != '' and row[8].value != '' \
            and 'Unavailable' not in row[4].value and 'Unavailable' not in row[7].value and row[4].value != '' and row[7].value != '':

        # similarity between the product names and product prices are calculated as a percentage(0-1)
        similarity = string_similar(row[4].value, row[7].value)
        price_ratio = float(row[5].value) / float(row[8].value)

        # ensure consistency in ratio(smaller price/larger price)
        if price_ratio < 1:
            price_ratio = 1 / price_ratio

        # new excel file will append all the data needed for the next step of comparison
        worksheet_2.append([similarity, row[4].value, row[7].value, row[5].value, row[8].value, price_ratio])
    else:
        worksheet_2.append([None, row[4].value, row[7].value, row[5].value, row[8].value, None])

# similarity thresholds are set for both prices and product names
Similarity_level = 0.5
p_ratio_level = 2

# rows from our new workbook excluding headings are used
rows_1 = worksheet_2.iter_rows(min_row=2)

# cells with possible errors are highlighted in red
fill_pattern = PatternFill(patternType='solid', fgColor='C64747')

# data from rows are cleaned
for row in rows_1:
    str(row[0].value).strip()

    # if there is valid data we can use
    if row[0].value != None and row[5].value != None and 'Unavailable' not in row[1].value and 'Unavailable' not in row[2].value \
            and row[3].value != '' and row[4].value != '':

        # if data is less than/greater than the similarity threshold set, the product cells are highlighted in red
        if float(row[0].value) < Similarity_level or row[5].value > p_ratio_level:
            row[1].fill = fill_pattern
            row[2].fill = fill_pattern

# new workbook saved under name
new_workbook.save('Data.xlsx')